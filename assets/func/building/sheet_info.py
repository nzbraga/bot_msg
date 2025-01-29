from openpyxl import load_workbook

def sheet_info():
    workbook = load_workbook('list_name.xlsx')
    sheet_name = workbook.sheetnames[0]
    sheet = workbook[sheet_name]

    return sheet