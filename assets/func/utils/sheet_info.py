from openpyxl import load_workbook
from assets.func.utils.pop_up import show_popup

def sheet_info(page):
    
    if page == -1:
        show_popup('Escolha uma pagina da planilha!')
        return False
    
    workbook = load_workbook('list_name.xlsx')
    sheet_name = workbook.sheetnames[page]
    sheet = workbook[sheet_name]

    return sheet